from tkinter import *
from tkinter.filedialog import askopenfilename
import tkinter.ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
import cx_Oracle as oracle

def selectPath():
    path_ = askopenfilename() #中包装调用文件选择框
    path.set(path_)

def selectPath1():
    path_ = askopenfilename() #菜子王小包装调用文件选择框
    path1.set(path_)

def selectPath2():
    path_ = askopenfilename() #道道全小包装调用文件选择框
    path2.set(path_)

def zbzghmbl():
    try:
        db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd')

        lujing = en1.get()

        wb = load_workbook(lujing)  # 文件名下
        #ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1
        ws = wb["Sheet1"]

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        lb101["text"] =""
        i = 0
        jdt['value'] = 0
        jdt['maximum'] = rows
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                #row_data.append(str(ws.cell(row=rx, column=cx).value))
                row_data.append(ws.cell(row=rx, column=cx).value)
            i = i+1
            jdt['value'] = i + 1
            sql = "UPDATE fr_zbzjxdc set ghmbl=:4 where 品牌=:1 and 日期=:2 and 区域=:3 and exists ( select 1 from (select 区域,日期,品牌,max(发货工厂编码) rm from fr_zbzjxdc where 品牌=:1  and 日期=:2 and 区域=:3 group by 区域,日期,品牌) a where 区域=fr_zbzjxdc.区域 and 日期=fr_zbzjxdc.日期 and rm=fr_zbzjxdc.发货工厂编码)"
            data = (row_data[3], row_data[0],row_data[1], row_data[2])

            print(data)
            cursor.execute(sql,data) # 执行sql语句
            row_data = []
            baseframe.update()
            lb6["text"] = i
            print(i)
            db.commit()
        cursor.close()
        db.close()# 关闭连接

        lb6["text"] = ""
        lb101["fg"] = "GREEN"
        lb101["text"] = '导入成功,共计%s条!'%i
    except Exception as e:
        print(e)
        lb101["fg"] = "red"
        lb101["text"] = '导入失败！'


def czwxbzghmbl():
    try:
        db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd')

        lujing = en11.get()

        wb = load_workbook(lujing)  # 文件名下
        #ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1
        ws = wb["Sheet1"]

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        lb101["text"] =""
        i = 0
        jdt['value'] = 0
        jdt['maximum'] = rows
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                #row_data.append(str(ws.cell(row=rx, column=cx).value))
                row_data.append(ws.cell(row=rx, column=cx).value)
            i = i+1
            jdt['value'] = i + 1
            sql = "UPDATE fr_czwxbzyjdc set ghmblxiao=:3,ghmbla=:4,ghmblxin=:5 where 日期=:1 and 区域=:2 and exists ( select 1 from (select 区域,日期,max(发货工厂编码) rm from fr_czwxbzyjdc where 日期=:1 and 区域=:2  group by 区域,日期) a where 区域=fr_czwxbzyjdc.区域 and 日期=fr_czwxbzyjdc.日期 and rm=fr_czwxbzyjdc.发货工厂编码)"
            data = (row_data[2],row_data[3],row_data[4],row_data[0],row_data[1])

            print(data)
            cursor.execute(sql,data) # 执行sql语句
            row_data = []
            baseframe.update()
            lb61["text"] = i
            print(i)
            db.commit()
        cursor.close()
        db.close()# 关闭连接

        lb61["text"] = ""
        lb101["fg"] = "GREEN"
        lb101["text"] = '导入成功,共计%s条!'%i
    except Exception as e:
        print(e)
        lb101["fg"] = "red"
        lb101["text"] = '导入失败！'


def ddqxbzghmbl():
    try:
        db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd')

        lujing = en12.get()

        wb = load_workbook(lujing)  # 文件名下
        #ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1
        ws = wb["Sheet1"]

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        lb101["text"] =""
        i = 0
        jdt['value'] = 0
        jdt['maximum'] = rows
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                #row_data.append(str(ws.cell(row=rx, column=cx).value))
                row_data.append(ws.cell(row=rx, column=cx).value)
            i = i+1
            jdt['value'] = i + 1
            sql = "UPDATE fr_ddqxbzyjdc set ghmblxiao=:3,ghmbla=:4,ghmblxin=:5,ghmblzs=:6,sqjl=:7,sqnq=:8,xsbjl=:9,xsbnq=:10,xsbch=:11 where 日期=:1 and 区域=:2"
            data = (row_data[2],row_data[3],row_data[4],row_data[5],row_data[6],row_data[7],row_data[8],row_data[9],row_data[10],row_data[0],row_data[1])

            print(data)
            cursor.execute(sql,data) # 执行sql语句

            row_data = []
            baseframe.update()
            lb62["text"] = i
            print(i)
            db.commit()

        sql2 = "update fr_ddqxbzyjdc set ghmblxiao='0',ghmbla='0',ghmblxin='0',ghmblzs='0' where exists ( select 1 from (select 区域,日期,max(发货工厂编码) rm from fr_ddqxbzyjdc group by 区域,日期 having count(*)>1) a where 区域=fr_ddqxbzyjdc.区域 and 日期=fr_ddqxbzyjdc.日期 and rm=fr_ddqxbzyjdc.发货工厂编码)"

        cursor.execute(sql2)
        db.commit()

        cursor.close()
        db.close()# 关闭连接

        lb62["text"] = ""
        lb101["fg"] = "GREEN"
        lb101["text"] = '导入成功,共计%s条!'%i
    except Exception as e:
        print(e)
        lb101["fg"] = "red"
        lb101["text"] = '导入失败！'


baseframe = Tk()
baseframe.wm_title("道道全营销中心规划目标量导入")
baseframe.geometry("800x600")

path = StringVar() #字符串变量”对象，可以与Entry、Label等控件绑定，这里的绑定是双向绑定，也就是既可以通过该变量来获取Entry、Label等控件中的值，也可以通过更改该变量来改变Entry、Label等控件中的值
path1 = StringVar()
path2 = StringVar()

photo = PhotoImage(file="bg.gif")#file：t图片路径

lb31 = Label(baseframe, text="", justify=LEFT, image=photo, compound=CENTER, font=("微软雅黑",16))
lb31.place(x=1, y=1, width=800, height=600)
#lb3.pack()
####################################中包装导入布局
lb0 = Label(baseframe,text="中包装规划目标量导入",bg="white",fg="blue",font=("微软雅黑", 18))
lb0.pack()
#lb0.place(x=140,y=2)

lb1 = Label(baseframe,text="目标路径:",bg="white",fg="red",font=("微软雅黑",12))
# lb1.pack()
lb1.place(x=145,y=30)

en1= Entry(baseframe, textvariable = path,width=55,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en1.place(x=220,y=33)

bt1 = Button(baseframe,text = "路径选择",command = selectPath,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt1.place(x=340,y=56)

bt2 = Button(baseframe,text = "导入",command = zbzghmbl,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt2.place(x=420,y=56)

lb6 = Label(baseframe,text="",bg="white",fg="green",font=("微软雅黑",15))
# lb1.pack()
lb6.place(x=615,y=32)

################################菜子王小包装布局
lb00 = Label(baseframe,text="菜子王小包装规划目标量导入",bg="white",fg="blue",font=("微软雅黑", 18))
#lb00.pack()
lb00.place(x=245,y=110)

lb11 = Label(baseframe,text="目标路径:",bg="white",fg="red",font=("微软雅黑",12))
# lb1.pack()
lb11.place(x=145,y=150)

en11= Entry(baseframe, textvariable = path1,width=55,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en11.place(x=220,y=150)

bt11 = Button(baseframe,text = "路径选择",command = selectPath1,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt11.place(x=340,y=173)

bt21 = Button(baseframe,text = "导入",command = czwxbzghmbl,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt21.place(x=420,y=173)

lb61 = Label(baseframe,text="",bg="white",fg="green",font=("微软雅黑",15))
# lb1.pack()
lb61.place(x=615,y=149)

############################道道全小包装布局
lb02 = Label(baseframe,text="道道全小包装规划目标量导入",bg="white",fg="blue",font=("微软雅黑", 18))
#lb00.pack()
lb02.place(x=245,y=220)

lb12 = Label(baseframe,text="目标路径:",bg="white",fg="red",font=("微软雅黑",12))
# lb1.pack()
lb12.place(x=145,y=255)

en12= Entry(baseframe, textvariable = path2,width=55,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en12.place(x=220,y=255)

bt12 = Button(baseframe,text = "路径选择",command = selectPath2,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt12.place(x=340,y=278)

bt22 = Button(baseframe,text = "导入",command = ddqxbzghmbl,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt22.place(x=420,y=278)

lb62 = Label(baseframe,text="",bg="white",fg="green",font=("微软雅黑",15))
# lb1.pack()
lb62.place(x=615,y=254)
####################################################################################

lb5 = Label(baseframe,text="导入进度:",bg="white",fg="red",font=("微软雅黑",12))
# lb1.pack()
lb5.place(x=150,y=450)

jdt = tkinter.ttk.Progressbar(baseframe,length=390)
jdt.place(x=225,y=450)

lb10 = tkinter.Label(baseframe, text="©Develop:Ike...Leo", bg="white", fg="green", font=("微软雅黑", 18))
# lb1.grid(row=0,column=0,stick=tkinter.W)
lb10.place(x=240, y=500)

lb101 = tkinter.Label(baseframe, text="", bg="white", fg="green", font=("微软雅黑", 20))
# lb1.grid(row=0,column=0,stick=tkinter.W)
lb101.place(x=250, y=400)

baseframe.mainloop()
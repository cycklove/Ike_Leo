from tkinter import *
from tkinter.filedialog import askopenfilename
import tkinter.ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
import cx_Oracle as oracle

def qingkongtable():
    db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd') #配置数据库连接
    cursor = db.cursor() #打开接收返回值方法
    sql = 'truncate table FR_OAYYZXZHIJI2' #需要执行的语句
    cursor.execute(sql)  #执行语句
    db.commit() #提交结果
    cursor.close() #关闭接收返回值方法
    db.close() #关闭数据库连接


def selectPath():
    path_ = askopenfilename() #调用文件选择框
    path.set(path_)


def insert_deta():
    qingkongtable()
    try:
        db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd')

        lujing = en1.get()

        wb = load_workbook(lujing)  # 文件名下
        #ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1
        ws = wb["Sheet1"]

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        lb3["text"] =""
        i = 0
        jdt['value'] = 0
        jdt['maximum'] = rows
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                #row_data.append(str(ws.cell(row=rx, column=cx).value))
                row_data.append(ws.cell(row=rx, column=cx).value)
            i = i+1
            jdt['value'] = i + 1
            sql = "INSERT INTO FR_OAYYZXZHIJI2(XM,YIJIBUMEN,ERJIBUMEN,SANJIBUMEN,QY,YF,ZHIWEI,ZHIJI,TIAOZHENGHOUZHIJI,GANGWEI,NF,SYNCTIME,PP,奖励标准小榨,奖励标准双低压榨菜籽油,占比系数,奖励标准,市场类别,five奖励标准)VALUES(:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,sysdate,:13,:14,:15,:16,:17,:18,:19)"
            data = (row_data[0], row_data[1],row_data[2], row_data[3],row_data[4], row_data[5],row_data[6], row_data[7],row_data[8], row_data[9],row_data[10],row_data[11],row_data[12],row_data[13],row_data[14],row_data[15],row_data[16],row_data[17])

            print(data)
            cursor.execute(sql,data) # 执行sql语句
            row_data = []
            baseframe.update()
            lb6["text"] = i
            print(i)
            db.commit()
        cursor.close()
        db.close()# 关闭连接

        lb6["text"] = ""
        lb3["fg"] = "GREEN"
        lb3["text"] = '导入成功,共计%s条!'%i
    except Exception as e:
        print(e)
        lb3["fg"] = "red"
        lb3["text"] = '导入失败！'


baseframe = Tk()
baseframe.wm_title("道道全营销中心职级导入")
baseframe.geometry("500x400")

path = StringVar() #字符串变量”对象，可以与Entry、Label等控件绑定，这里的绑定是双向绑定，也就是既可以通过该变量来获取Entry、Label等控件中的值，也可以通过更改该变量来改变Entry、Label等控件中的值

photo = PhotoImage(file="bg.gif")#file：t图片路径

lb3 = Label(baseframe,text = "",justify = LEFT,image = photo,compound = CENTER,font=("微软雅黑",28))
lb3.place(x=1,y=-150,width=500,height=600)
#lb3.pack()

lb0 = Label(baseframe,text="营销职级导入",bg="white",fg="blue",font=("微软雅黑", 18))
lb0.pack()
#lb0.place(x=140,y=2)

lb1 = Label(baseframe,text="目标路径:",bg="white",fg="black",font=("微软雅黑",12))
# lb1.pack()
lb1.place(x=1,y=30)


en1= Entry(baseframe, textvariable = path,width=59,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en1.place(x=76,y=33)

bt1 = Button(baseframe,text = "路径选择",command = selectPath,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt1.place(x=170,y=56)

bt2 = Button(baseframe,text = "导入",command = insert_deta,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt2.place(x=250,y=56)

lb4 = Label(baseframe, text="©Develop:Ike...Leo", bg="white", fg="green", font=("微软雅黑", 18))
# lb1.grid(row=0,column=0,stick=tkinter.W)
lb4.place(x=135,y=305)

lb5 = Label(baseframe,text="导入进度:",bg="white",fg="black",font=("微软雅黑",12))
# lb1.pack()
lb5.place(x=60,y=100)

jdt = tkinter.ttk.Progressbar(baseframe,length=250)
jdt.place(x=135,y=100)

lb6 = Label(baseframe,text="",bg="white",fg="green",font=("微软雅黑",15))
# lb1.pack()
lb6.place(x=385,y=95)

baseframe.mainloop()

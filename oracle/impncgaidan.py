from tkinter import *
from tkinter.filedialog import askopenfilename
import tkinter.ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
import cx_Oracle as oracle


def qingkongtable():
    db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd') #配置数据库连接
    cursor = db.cursor() #打开接收返回值方法
    sql = 'truncate table fr_nc_gaidan_temp' #需要执行的语句
    cursor.execute(sql)  #执行语句
    db.commit() #提交结果
    cursor.close() #关闭接收返回值方法
    db.close() #关闭数据库连接


def selectPath():
    path_ = askopenfilename() #调用文件选择框
    path.set(path_)


def insert_deta():
    nr = en1.get()
    if '.xlsx' in nr:
        qingkongtable()
    try:
        db = oracle.connect('ddqnc63/ddqnc63@192.168.50.31:1521/ncprd')

        lujing = en1.get()

        wb = load_workbook(lujing)  # 文件名下
        #ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1
        ws = wb["Sheet1"]

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        lb3["text"] =""
        i = 0
        jdt['value'] = 0
        jdt['maximum'] = rows
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                #row_data.append(str(ws.cell(row=rx, column=cx).value))
                row_data.append(ws.cell(row=rx, column=cx).value)
            i = i+1
            jdt['value'] = i + 1
            sql = "INSERT INTO fr_nc_gaidan_temp(NC订单号,物料编码,订单行数量,年度返利总金额,目标达成总金额,新物料模式总金额,市场费用总金额,新品奖励总金额,其他摊销总金额,开票价,月度优惠单价,特价优惠单价,限量优惠单价,赠品摊销,其他政策优惠,内供价,小油是否摊销)VALUES(:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16,:17)"
            data = (row_data[0], row_data[1],row_data[2], row_data[3],row_data[4], row_data[5],row_data[6], row_data[7],row_data[8], row_data[9],row_data[10],row_data[11],row_data[12],row_data[13],row_data[14],row_data[15],row_data[16])

            print(data)
            cursor.execute(sql,data) # 执行sql语句
            row_data = []
            baseframe.update()
            lb6["text"] = i
            print(i)
            db.commit()

        sqlqian = "select count(1) from fr_nc_gaidan"
        cursor.execute(sqlqian)  # 执行sql语句
        row = cursor.fetchone()  # 读取查询结果,
        while row:  # 循环读取所有结果
            iqian = row[0]
            print(iqian)  # 输出结果
            row = cursor.fetchone()

        sql2 = "MERGE INTO fr_nc_gaidan A USING (select B.NC订单号,B.物料编码,B.订单行数量,B.年度返利总金额,B.目标达成总金额,B.新物料模式总金额,B.市场费用总金额,B.新品奖励总金额,B.其他摊销总金额,B.开票价,B.月度优惠单价,B.特价优惠单价,B.限量优惠单价,B.赠品摊销,B.其他政策优惠, B.内供价,B.小油是否摊销 from fr_nc_gaidan_temp B) C \
ON (A.NC订单号=C.NC订单号 and A.物料编码=C.物料编码 and A.订单行数量=C.订单行数量)\
WHEN MATCHED THEN \
UPDATE SET A.年度返利总金额=C.年度返利总金额,A.目标达成总金额=C.目标达成总金额,A.新物料模式总金额=C.新物料模式总金额,A.市场费用总金额=C.市场费用总金额,A.新品奖励总金额=C.新品奖励总金额,A.其他摊销总金额=C.其他摊销总金额,\
A.开票价=C.开票价,A.月度优惠单价=C.月度优惠单价,A.特价优惠单价=C.特价优惠单价,A.限量优惠单价=C.限量优惠单价,A.赠品摊销=C.赠品摊销,A.其他政策优惠=C.其他政策优惠,A.内供价=C.内供价,A.小油是否摊销=C.小油是否摊销 \
WHEN NOT MATCHED THEN\
 INSERT(A.NC订单号,A.物料编码,A.订单行数量,A.年度返利总金额,A.目标达成总金额,A.新物料模式总金额,A.市场费用总金额,A.新品奖励总金额,A.其他摊销总金额,A.开票价,A.月度优惠单价,A.特价优惠单价,A.限量优惠单价,A.赠品摊销,A.其他政策优惠,A.内供价,A.小油是否摊销) \
VALUES(C.NC订单号,C.物料编码,C.订单行数量,C.年度返利总金额,C.目标达成总金额,C.新物料模式总金额,C.市场费用总金额,C.新品奖励总金额,C.其他摊销总金额,C.开票价,C.月度优惠单价,C.特价优惠单价,C.限量优惠单价,C.赠品摊销,C.其他政策优惠,C.内供价,C.小油是否摊销)"
        cursor.execute(sql2)
        db.commit()

        sqlhou = "select count(1) from fr_nc_gaidan"
        cursor.execute(sqlhou)  # 执行sql语句
        row = cursor.fetchone()  # 读取查询结果,
        while row:  # 循环读取所有结果
            ihou = row[0]
            print(ihou)  # 输出结果
            row = cursor.fetchone()

        cursor.close()
        db.close()# 关闭连接

        idiff = int(ihou) - int(iqian)
        iup = int(i) - int(idiff)

        lb6["text"] = ""
        lb3["fg"] = "GREEN"
        txt1 = '导入成功,共计%s条!'%i
        txt2 = '其中新增%s条!'%idiff
        txt3 = '修改%s条!' %iup
        lb3["text"] = txt1+txt2+txt3
    except Exception as e:
        print(e)
        lb3["fg"] = "red"
        lb3["text"] = '导入失败！'


baseframe = Tk()
baseframe.wm_title("财务NC改单导入")
baseframe.geometry("500x400")

path = StringVar() #字符串变量”对象，可以与Entry、Label等控件绑定，这里的绑定是双向绑定，也就是既可以通过该变量来获取Entry、Label等控件中的值，也可以通过更改该变量来改变Entry、Label等控件中的值

photo = PhotoImage(file="bg.gif")#file：t图片路径

lb3 = Label(baseframe,text = "",justify = LEFT,image = photo,compound = CENTER,font=("微软雅黑",16))
lb3.place(x=1,y=-150,width=500,height=600)
#lb3.pack()

lb0 = Label(baseframe,text="财务NC改单导入",bg="white",fg="blue",font=("微软雅黑", 18))
lb0.pack()
#lb0.place(x=140,y=2)

lb1 = Label(baseframe,text="目标路径:",bg="white",fg="black",font=("微软雅黑",12))
# lb1.pack()
lb1.place(x=1,y=30)


en1= Entry(baseframe, textvariable = path,width=59,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en1.place(x=76,y=33)

bt1 = Button(baseframe,text = "路径选择",command = selectPath,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt1.place(x=170,y=56)

bt2 = Button(baseframe,text = "导入",command = insert_deta,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt2.place(x=250,y=56)

lb4 = Label(baseframe, text="©Develop:Ike...Leo", bg="white", fg="green", font=("微软雅黑", 18))
# lb1.grid(row=0,column=0,stick=tkinter.W)
lb4.place(x=135,y=305)

lb5 = Label(baseframe,text="导入进度:",bg="white",fg="black",font=("微软雅黑",12))
# lb1.pack()
lb5.place(x=60,y=100)

jdt = tkinter.ttk.Progressbar(baseframe,length=250)
jdt.place(x=135,y=100)

lb6 = Label(baseframe,text="",bg="white",fg="green",font=("微软雅黑",15))
# lb1.pack()
lb6.place(x=385,y=95)

baseframe.mainloop()

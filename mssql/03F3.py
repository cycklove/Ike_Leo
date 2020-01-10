import pymssql
from itertools import chain
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment

def apossaleswenjian():
    dbname = ['FRXS_ERP_ORDER_200', 'FRXS_ERP_ORDER_204', 'FRXS_ERP_ORDER_208', 'FRXS_ERP_ORDER_212',
'FRXS_ERP_ORDER_216','FRXS_ERP_ORDER_220', 'FRXS_ERP_ORDER_224', 'FRXS_ERP_ORDER_228', 'FRXS_ERP_ORDER_232',
'FRXS_ERP_ORDER_238','FRXS_ERP_ORDER_250', 'FRXS_ERP_ORDER_255', 'FRXS_ERP_ORDER_265', 'FRXS_ERP_ORDER_270',
'FRXS_ERP_ORDER_275','FRXS_ERP_ORDER_280', 'FRXS_ERP_ORDER_285', 'FRXS_ERP_ORDER_290', 'FRXS_ERP_ORDER_295',
'FRXS_ERP_ORDER_300', 'FRXS_ERP_ORDER_305', 'FRXS_ERP_ORDER_310', 'FRXS_ERP_ORDER_315','FRXS_ERP_ORDER_320',
'FRXS_ERP_ORDER_325', 'FRXS_ERP_ORDER_330', 'FRXS_ERP_ORDER_335', ]

    for i in dbname:
        db = pymssql.connect("192.168.3.141", "sa", "frxs#@!admin", i)
        c = db.cursor()
        with open('test.sql',encoding='utf-8',mode='r') as f:
            # 读取整个sql文件，以分号切割。[:-1]删除最后一个元素，也就是空字符串
            sql_list = f.read().split(';')[:-1]
            for x in sql_list:
                # 判断包含空行的
                if '\n' in x:
                    # 替换空行为1个空格
                    x = x.replace('\n', ' ')
                # 判断多个空格时
                if '    ' in x:
                    # 替换为空
                    x = x.replace('    ','')
                # sql语句添加分号结尾
                sql_item = x+';'
                # print(sql_item)
                c.execute(sql_item)
                print("执行成功sql: %s"%sql_item)

                c.close()
                db.commit()
                db.close()

def apossalesjiaoben():

    dbname = ['FRXS_ERP_ORDER_200', 'FRXS_ERP_ORDER_204', 'FRXS_ERP_ORDER_208', 'FRXS_ERP_ORDER_212','FRXS_ERP_ORDER_216',
'FRXS_ERP_ORDER_220', 'FRXS_ERP_ORDER_224', 'FRXS_ERP_ORDER_228', 'FRXS_ERP_ORDER_232','FRXS_ERP_ORDER_238',
'FRXS_ERP_ORDER_250', 'FRXS_ERP_ORDER_255', 'FRXS_ERP_ORDER_265', 'FRXS_ERP_ORDER_270','FRXS_ERP_ORDER_275',
'FRXS_ERP_ORDER_280', 'FRXS_ERP_ORDER_285', 'FRXS_ERP_ORDER_290', 'FRXS_ERP_ORDER_295',
'FRXS_ERP_ORDER_300', 'FRXS_ERP_ORDER_305', 'FRXS_ERP_ORDER_310', 'FRXS_ERP_ORDER_315',
'FRXS_ERP_ORDER_320','FRXS_ERP_ORDER_325', 'FRXS_ERP_ORDER_330', 'FRXS_ERP_ORDER_335', ]

    for i in dbname:
        db = pymssql.connect("192.168.3.141", "sa", "frxs#@!admin", i)
        c = db.cursor()
        sql_item = "delete from c_ike where id = 3;"
        c.execute(sql_item)
        print("执行成功sql: %s"%sql_item)
        db.commit()
        c.close()
        db.close()


def apossalesselect():

    dbname = ['FRXS_ERP_ORDER_200', 'FRXS_ERP_ORDER_204', 'FRXS_ERP_ORDER_208', 'FRXS_ERP_ORDER_212','FRXS_ERP_ORDER_216',
'FRXS_ERP_ORDER_220', 'FRXS_ERP_ORDER_224', 'FRXS_ERP_ORDER_228', 'FRXS_ERP_ORDER_232','FRXS_ERP_ORDER_238',
'FRXS_ERP_ORDER_250', 'FRXS_ERP_ORDER_255', 'FRXS_ERP_ORDER_265', 'FRXS_ERP_ORDER_270','FRXS_ERP_ORDER_275',
'FRXS_ERP_ORDER_280', 'FRXS_ERP_ORDER_285', 'FRXS_ERP_ORDER_290', 'FRXS_ERP_ORDER_295',
'FRXS_ERP_ORDER_300','FRXS_ERP_ORDER_305', 'FRXS_ERP_ORDER_310', 'FRXS_ERP_ORDER_315', 'FRXS_ERP_ORDER_320',
'FRXS_ERP_ORDER_325','FRXS_ERP_ORDER_330', 'FRXS_ERP_ORDER_335', ]

    for i in dbname:
        db = pymssql.connect("192.168.3.141", "sa", "frxs#@!admin", i)
        c = db.cursor()
        sql_item = 'select id from c_ike'
        c.execute(sql_item)
        # row = c.fetchall()
        # for i in row:
        #     print(i)
        row = c.fetchone()
        while row:
            print(i,row[0])
            row = c.fetchone()
        # print("执行成功sql: %s"%sql_item)
        db.commit()
        c.close()
        db.close()

def apossalestable():

    dbname=['FRXS_ERP_ORDER_200','FRXS_ERP_ORDER_204','FRXS_ERP_ORDER_208','FRXS_ERP_ORDER_212','FRXS_ERP_ORDER_216',
'FRXS_ERP_ORDER_220','FRXS_ERP_ORDER_224','FRXS_ERP_ORDER_228','FRXS_ERP_ORDER_232','FRXS_ERP_ORDER_238',
'FRXS_ERP_ORDER_250','FRXS_ERP_ORDER_255','FRXS_ERP_ORDER_265','FRXS_ERP_ORDER_270','FRXS_ERP_ORDER_275',
'FRXS_ERP_ORDER_280','FRXS_ERP_ORDER_285','FRXS_ERP_ORDER_290','FRXS_ERP_ORDER_295','FRXS_ERP_ORDER_300',
'FRXS_ERP_ORDER_305','FRXS_ERP_ORDER_310','FRXS_ERP_ORDER_315','FRXS_ERP_ORDER_320','FRXS_ERP_ORDER_325',
'FRXS_ERP_ORDER_330','FRXS_ERP_ORDER_335']

    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    iname = "fyxs"

    sheetname = iname
    ws = wb.create_sheet(iname)
    ws = wb.get_sheet_by_name(iname)
    ws['A1'] = '仓库'
    ws['B1'] = '1月'
    ws['C1'] = '2月'
    ws['D1'] = '3月'
    ws['E1'] = '4月'
    ws['F1'] = '5月'
    ws['G1'] = '6月'
    ws['H1'] = '7月'
    ws['I1'] = '8月'
    ws['J1'] = '9月'
    ws['K1'] = '10月'
    ws['L1'] = '11月'
    ws['M1'] = '12月'

    bold_itatic_24_font = Font(bold=True)
    ws['A1'].font = bold_itatic_24_font
    ws['B1'].font = bold_itatic_24_font
    ws['C1'].font = bold_itatic_24_font
    ws['D1'].font = bold_itatic_24_font
    ws['E1'].font = bold_itatic_24_font
    ws['F1'].font = bold_itatic_24_font
    ws['G1'].font = bold_itatic_24_font
    ws['H1'].font = bold_itatic_24_font
    ws['I1'].font = bold_itatic_24_font
    ws['J1'].font = bold_itatic_24_font
    ws['K1'].font = bold_itatic_24_font
    ws['L1'].font = bold_itatic_24_font
    ws['M1'].font = bold_itatic_24_font

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

    for i in dbname:
        db = pymssql.connect("192.168.3.141", "sa", "frxs#@!admin", i)
        c = db.cursor()

        sql_item = "select d.WName,sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201901 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201902 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201903 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201904 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201905 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201906 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201907 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201908 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201909 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201910 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201911 then cb.xs+cb.th end),0)),sum(isnull((case when convert(varchar(6),cb.sett_date,112) = 201912 then cb.xs+cb.th end),0)) from (select a.WName,a.WID,a.sett_date,b.ProductId,b.SubAmt as xs,b.UnitQty as xssl,0 as th,0 as thsl from vSaleOrder a with (nolock) inner join vSaleOrderDetails b with (nolock) on a.OrderId = b.OrderID where a.Sett_Date between '2019-01-01' and '2019-12-31' union all select a.WName,a.WID,a.sett_date,b.ProductId,0 as xs,0 AS xssl,-b.SubAmt as th,-b.UnitQty as thsl from SaleBack a with (nolock) inner join SaleBackDetails b with (nolock) on a.BackID = b.BackID where a.Sett_Date between '2019-01-01' and '2019-12-31') cb inner join frxs_erp_basedata.dbo.Products a on cb.ProductId = a.ProductId inner join frxs_erp_basedata.dbo.BaseProducts b on cb.ProductId = b.BaseProductId inner join frxs_erp_basedata.dbo.Categories c on b.CategoryId3=c.CategoryId inner join FRXS_ERP_BASEDATA.dbo.Warehouse d on cb.WID = d.wid where CHARINDEX('奖',c.Name)=0 group by d.WName"
        
        c.execute(sql_item)
        # row = c.fetchall()
        # print(i,sql_item,row)
        row = c.fetchone() #

        # if row: # 按查询结果分多个sheet保存
        #     iname = row[0]
        #
        #     sheetname = iname
        #     ws = wb.create_sheet(iname)
        #     ws = wb.get_sheet_by_name(iname)
        #     ws['A1'] = '仓库'
        #     ws['B1'] = '1月'
        #     ws['C1'] = '2月'
        #     ws['D1'] = '3月'
        #     ws['E1'] = '4月'
        #     ws['F1'] = '5月'
        #     ws['G1'] = '6月'
        #     ws['H1'] = '7月'
        #     ws['I1'] = '8月'
        #     ws['J1'] = '9月'
        #     ws['K1'] = '10月'
        #     ws['L1'] = '11月'
        #     ws['M1'] = '12月'
        #
        #     bold_itatic_24_font = Font(bold=True)
        #     ws['A1'].font = bold_itatic_24_font
        #     ws['B1'].font = bold_itatic_24_font
        #     ws['C1'].font = bold_itatic_24_font
        #     ws['D1'].font = bold_itatic_24_font
        #     ws['E1'].font = bold_itatic_24_font
        #     ws['F1'].font = bold_itatic_24_font
        #     ws['G1'].font = bold_itatic_24_font
        #     ws['H1'].font = bold_itatic_24_font
        #     ws['I1'].font = bold_itatic_24_font
        #     ws['J1'].font = bold_itatic_24_font
        #     ws['K1'].font = bold_itatic_24_font
        #     ws['L1'].font = bold_itatic_24_font
        #
        #     ws.column_dimensions['A'].width = 15
        #     ws.column_dimensions['B'].width = 15
        #     ws.column_dimensions['C'].width = 15
        #     ws.column_dimensions['D'].width = 15
        #     ws.column_dimensions['E'].width = 15
        #     ws.column_dimensions['F'].width = 15
        #     ws.column_dimensions['G'].width = 15
        #     ws.column_dimensions['H'].width = 15
        #     ws.column_dimensions['I'].width = 15
        #     ws.column_dimensions['J'].width = 15
        #     ws.column_dimensions['K'].width = 15
        #     ws.column_dimensions['L'].width = 15
        #     ws.column_dimensions['M'].width = 15

        while row: # 使用while和for循环只会返回有值的语句信息
            print(row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12])
            ws.append([row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12]])
            row = c.fetchone()

    wb.create_sheet("Sheet")
    wb.save('fyxs.xlsx')
    db.commit()
    c.close()
    db.close()

if __name__ == '__main__':
    # apossaleswenjian()
    # apossalesjiaoben()
    # apossalesselect()
    apossalestable()
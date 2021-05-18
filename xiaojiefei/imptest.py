from tkinter import *
from tkinter.filedialog import askopenfilename
import pymssql
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

def selectPath():
    path_ = askopenfilename()
    path.set(path_)


def insert_deta():
    try:
        db = pymssql.connect("192.168.10.84", "sa", "AAAaaa123", "cb","utf8")

        lujing = en1.get()

        wb = load_workbook(lujing)  # 文件名下
        ws = wb.get_sheet_by_name("Sheet1")  # execl里面的worksheet1

        cursor = db.cursor()

        rows = ws.max_row  # 最大行数
        columns = ws.max_column  # 最大列数

        row_data = []
        for rx in range(2, rows + 1):
            for cx in range(1, columns + 1):
                row_data.append(str(ws.cell(row=rx, column=cx).value))
            sql = "INSERT INTO tb002(bm,mc )VALUES(%s,%s)"
            cursor.execute(sql,(row_data[0], row_data[1])) # 执行sql语句
            row_data = []

            db.commit()
        cursor.close()
        db.close()# 关闭连接

        lb3["fg"] = "GREEN"
        lb3["text"] = '导入成功！'
    except:
        lb3["fg"] = "red"
        lb3["text"] = '导入失败！'


baseframe = Tk()
baseframe.wm_title("小劫匪F3门店导入专用")
baseframe.geometry("500x400")

path = StringVar()

photo = PhotoImage(file="D:\PycharmProjects\Ike_Leo\爬虫\\bg.gif")#file：t图片路径

lb3 = Label(baseframe,text = "",justify = LEFT,image = photo,compound = CENTER,font=("微软雅黑",28))
lb3.place(x=1,y=1,width=500,height=600)

lb0 = Label(baseframe,text="杨涵之小劫匪F3门店导入工具",bg="white",fg="blue",font=("微软雅黑",12))
# lb1.pack()
lb0.place(x=140,y=2)

lb1 = Label(baseframe,text="目标路径:",bg="white",fg="black",font=("微软雅黑",12))
# lb1.pack()
lb1.place(x=1,y=30)


en1= Entry(baseframe, textvariable = path,width=59,insertbackground='black', highlightthickness=1,fg="green")
# en1.pack()
en1.place(x=76,y=33)

bt1 = Button(baseframe,text = "路径选择",command = selectPath,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt1.pack()
bt1.place(x=170,y=56)

bt2 = Button(baseframe,text = "导入",command = insert_deta,fg="black",font=("微软雅黑",12),highlightthickness=1)
# bt2.pack()
bt2.place(x=250,y=56)

baseframe.mainloop()



# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Alignment

def chuli():

    tbname = r"D:\5008.xlsx"

    wb = load_workbook(tbname)  # 文件名下

    ws = wb["Sheet1"]

    rows = ws.max_row  # 最大行数
    columns = ws.max_column  # 最大列数

    maxvl = ws.cell(row=rows, column=1).value
    print(maxvl)
    row_data = []
    for rx in range(2, rows + 1):
        row_data.append([ws.cell(row=rx, column=1).value, ws.cell(row=rx, column=2).value])
    #print(row_data)

    row_data2 = []
    i = 0

    while i < maxvl:
        i = i+0.001
        # print(round(i,3))
        row_data2.append([round(i, 3), 0])

    dataind = len(row_data)

    row_data3 = []
    i2 = 0
    while i2 < dataind:
        row_data3.append(row_data[i2][0])
        i2 = i2+1

    print(dataind,len(row_data3))

    for b in row_data2:
        if b[0] in row_data3:
            pass
        else:
            row_data.append(b)

    print(len(row_data))
    row_data.sort()
    for c in row_data:
        # print(c,row_data.index(c))
        if c[1] == 0:
            # print(c,round((row_data[row_data.index(c)-1][1]+row_data[row_data.index(c)+1][1])/2,3))
            up = [c[0], round((row_data[row_data.index(c) - 1][1] + row_data[row_data.index(c) + 1][1]) / 2, 3)]
            row_data[row_data.index(c)] = up

    # wb2 = Workbook()
    # wb2.remove_sheet(wb2.get_sheet_by_name("Sheet2"))
    # wb2.remove_sheet(wb2.get_sheet_by_name("Sheet3"))
    shname = "solved"
    wb.create_sheet(shname, 1)
    ws2 = wb[shname]
    ws2['A1'] = '高度(m)'
    ws2['B1'] = '容积(L)'

    bold_itatic_24_font = Font(bold=True)
    ws2['A1'].font = bold_itatic_24_font
    ws2['B1'].font = bold_itatic_24_font

    for d in row_data:
        print(d)
        ws2.append([d[0],d[1]])
    wb.save(tbname)
if __name__ == '__main__':
    chuli()
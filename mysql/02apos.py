import pymysql
from itertools import chain
from openpyxl import Workbook

def apossaleswenjian():

    dbname=['apos_sales_0','apos_sales_1','apos_sales_10','apos_sales_11','apos_sales_12','apos_sales_13','apos_sales_14',
    'apos_sales_15','apos_sales_16','apos_sales_17','apos_sales_18','apos_sales_19','apos_sales_2','apos_sales_3',
    'apos_sales_4','apos_sales_5','apos_sales_6','apos_sales_7','apos_sales_8','apos_sales_9']

    for i in dbname:
        db = pymysql.connect("47.112.168.44", "root", "Al8!6s&ljaf20L", i)
        c = db.cursor()
        with open('test.sql',encoding='utf-8',mode='r') as f:
            # 读取整个sql文件，以分号切割。[:-1]删除最后一个元素，也就是空字符串
            sql_list = f.read().split(';')[:-1]
            for x in sql_list:
                # 判断包含空行的
                if '\n' in x:
                    # 替换空行为1个空格
                    x = x.replace('\n', ' ')
                # 判断多个空格时
                if '    ' in x:
                    # 替换为空
                    x = x.replace('    ','')
                # sql语句添加分号结尾
                sql_item = x+';'
                # print(sql_item)
                c.execute(sql_item)
                print("执行成功sql: %s"%sql_item)

                c.close()
                db.commit()
                db.close()

def apossalesjiaoben():

    dbname=['apos_sales_0','apos_sales_1','apos_sales_10','apos_sales_11','apos_sales_12','apos_sales_13','apos_sales_14',
    'apos_sales_15','apos_sales_16','apos_sales_17','apos_sales_18','apos_sales_19','apos_sales_2','apos_sales_3',
    'apos_sales_4','apos_sales_5','apos_sales_6','apos_sales_7','apos_sales_8','apos_sales_9']

    for i in dbname:
        db = pymysql.connect("47.112.168.44", "root", "Al8!6s&ljaf20L", i)
        c = db.cursor()
        sql_item = "delete from c_ike where id = 3;"
        c.execute(sql_item)
        print("执行成功sql: %s"%sql_item)
        db.commit()
        c.close()
        db.close()


def apossalesselect():

    dbname=['apos_sales_0','apos_sales_1','apos_sales_10','apos_sales_11','apos_sales_12','apos_sales_13','apos_sales_14',
    'apos_sales_15','apos_sales_16','apos_sales_17','apos_sales_18','apos_sales_19','apos_sales_2','apos_sales_3',
    'apos_sales_4','apos_sales_5','apos_sales_6','apos_sales_7','apos_sales_8','apos_sales_9']

    for i in dbname:
        db = pymysql.connect("47.112.168.44", "root", "Al8!6s&ljaf20L", i)
        c = db.cursor()
        sql_item = 'select id from c_ike'
        c.execute(sql_item)
        # row = c.fetchall()
        # for i in row:
        #     print(i)
        row = c.fetchone()
        while row:
            print(i,row[0])
            row = c.fetchone()
        # print("执行成功sql: %s"%sql_item)
        db.commit()
        c.close()
        db.close()

def apossalestable():

    dbname=['apos_sales_0','apos_sales_1','apos_sales_10','apos_sales_11','apos_sales_12','apos_sales_13','apos_sales_14',
    'apos_sales_15','apos_sales_16','apos_sales_17','apos_sales_18','apos_sales_19','apos_sales_2','apos_sales_3',
    'apos_sales_4','apos_sales_5','apos_sales_6','apos_sales_7','apos_sales_8','apos_sales_9']

    dbname2 = ['apos_goods_0', 'apos_goods_1']

    wb = Workbook()
    sheetname = 'Ike_Data'
    ws = wb['Sheet']
    ws.title = sheetname
    ws = wb.get_sheet_by_name(sheetname)
    ws['A1'] = '门店编码'
    ws['B1'] = '条数'

    for i in dbname2:
        db = pymysql.connect("47.112.168.44", "root", "Al8!6s&ljaf20L", i)
        c = db.cursor()
        for itb in range(0,100):
            sql_item = "select storeSysCode,count(1) from pos_item_"+str(itb)+" where storeSysCode in ('60816','60795','60797','60865','60863','60865','60880','60877','60870','60876','60892','60944','60839','60870','60438','60517','60827','60796', '60787','60754','60777','60757','60824','60816','60820','60837','60842','60851','60794','60749','60897','60489') group by storeSysCode;"
            c.execute(sql_item)
            # row = c.fetchall()
            # print(i,sql_item,row)
            row = c.fetchone() #
            while row: # 使用while和for循环只会返回有值的语句信息
                print(row[0],row[1])
                ws.append([row[0],row[1]])
                row = c.fetchone()
        wb.save('miandiantiaoshu.xlsx')
        db.commit()
        c.close()
        db.close()

if __name__ == '__main__':
    # apossaleswenjian()
    # apossalesjiaoben()
    # apossalesselect()
    apossalestable()